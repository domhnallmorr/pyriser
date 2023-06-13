import os

import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from wood_dw.misc import openpyxl_custom_styles

class TabFile():
	"""
	Author:
	-------------
		- Domhnall Morrisey 26/05/2021
		
	Prerequisites:
	-------------
		- Python > 3.9
		- openpyxl

	Limitations: 
	-------------
		- Limited error handling at the moment

	Key Variables: 
	-------------
		- self.tab_file_data (dict containing the data from each table in the tab file)
		
	TO DO: 
	-------------
		- Improve verbose and error handling
		- Might remove table number in tab_file_data keys
		- Might add some functionality to utilise das file data (e.g. finding which joint contains the max VMS)
	"""
	def __init__(self, tab_file, contents=None, print_to_screen=False):

		'''
		:param tab_file: (``string``) - full path of tab file file to be read
		:param contents: (``list`` - optional) - tab file line data. Normally should be None, so the script will read the data from the file.
									Honestly can't remember why I included this. Must have been some edge case where it was required.
		:param print_to_screen (``boolean`` - optional): - Print file name if True (default value True). Useful for tracking progress if reading multiple files
		'''	
		
		assert os.path.isfile(tab_file), f'Tab file {tab_file} does not exist!'
		
		self.contents = contents
		self.tab_file = tab_file
		self.print_to_screen = print_to_screen
		
		if not contents:
			f = open(tab_file)
			
			self.contents = f.readlines()
			
			f.close()
			
		self.tab_file_data = {}
		self.parse_contents()
		
		self.process_automated_postprocessing()
			
	def parse_contents(self):
		if self.print_to_screen:
			print('Parsing Tab File Contents')
		table = None
		
		for idx, line in enumerate(self.contents):	
		
			if 'Table No.' in line:
				table = line.replace('\n', '')

				self.tab_file_data[table] = {}
				series_count = 0
				x_axis_label = ''
				y_axis_label = ''
				x_units = ''
				y_units = ''
				
			elif table:

				if line[0:7] == 'X-Axis:':
					x_axis_label = line.split(':')[1].strip()
				
				if line[0:7] == 'Y-Axis:':
					y_axis_label = line.split(':')[1].strip()
					
				if line[0:8] == 'X-Units:':
					x_units = line.split(':')[1].strip()
					
					if x_units != '':
						x_axis_label = f'{x_axis_label} ({x_units})'
						
				if line[0:8] == 'Y-Units:':
					y_units = line.split(':')[1].strip()			

					if y_units != '':
						y_axis_label = f'{y_axis_label} ({y_units})'
						
				if line[0:10] == 'Plot Data:':
					series = line.split(':')[1].strip()
					
					if series == '' or series in self.tab_file_data[table].keys(): # if not plot data title is defined
						series_count += 1
						series = f'Series {series_count}'
					
					self.tab_file_data[table][series] = {'X': [], 'Y': [], 'X Label': x_axis_label, 'Y Label': y_axis_label, 'X Units': x_units, 'Y Units': y_units}
					
				try:
					line = line.split('\t')
					self.tab_file_data[table][series]['X'].append(float(line[0]))
					self.tab_file_data[table][series]['Y'].append(float(line[1]))
				except:
					pass			
		
	def process_automated_postprocessing(self):
		
		self.max_vms = None
		self.max_vms_top = None
		self.max_vms_btm = None

		self.max_bm = None
		self.max_bm_btm = None
		self.max_bm_top = None
		self.riser_length = None
		
		table_map = {'Envelope of Von Mises Stress': {'series': 'Max-Von Mises Stress Envelope', 'top': self.max_vms_top, 'btm': self.max_vms_btm}, 
			'Envelope of Resultant Bending Moment': {'series': 'Max-Resultant Bending Moment Envelope', 'top': self.max_bm_top, 'btm': self.max_bm_btm},
				}

		for table in self.tab_file_data.keys():
			table_name = table.split(' - ')[1].strip()
			for param in table_map.keys():
				if param == table_name:
					series = table_map[param]['series']
					if series in self.tab_file_data[table].keys():

						mid_idx = int(len(self.tab_file_data[table][series]['Y'])/2)

						self.set_value(series, table, mid_idx)
						
						if not self.riser_length:
							self.riser_length = max(self.tab_file_data[table][series]['Y']) -  min(self.tab_file_data[table][series]['Y'])
				
					
	def set_value(self, series, table, mid_idx):

		if series == 'Max-Von Mises Stress Envelope':
			
			self.max_vms_top = max(self.tab_file_data[table][series]['X'][:mid_idx])
			self.max_vms_btm = max(self.tab_file_data[table][series]['X'][mid_idx:])

		if series == 'Max-Resultant Bending Moment Envelope':
			
			self.max_bm_top = max(self.tab_file_data[table][series]['X'][:mid_idx])
			self.max_bm_btm = max(self.tab_file_data[table][series]['X'][mid_idx:])
			
	def export_to_excel(self, excel_file, to_specific_tab=False):
	
		wb = Workbook()
		excel_styler = openpyxl_custom_styles.OpenpyxlCustomStyles()
		
		if not to_specific_tab:
			wb.create_sheet('Tab Data')
			wb.active = wb['Tab Data']
			sheet = wb.active
		
		col = 1
		for table in self.tab_file_data.keys():
			row = 3
			excel_styler.write_header_row(sheet, [table], col, row-2)
			
			for series in self.tab_file_data[table]:
				excel_styler.write_header_row(sheet, [series], col, row-1)
				
				data = []
				for idx, x in enumerate(self.tab_file_data[table][series]["X"]):
					data.append([x, self.tab_file_data[table][series]["Y"][idx]])
				
				excel_styler.write_tabular_data(sheet, data, col, row)
				col += 2
				
			
		wb.save("test.xlsx")
		
	def get_table_number(self, table):
		
		table_number = int(table.split(' ')[1].replace('No.', ''))
	
		return table_number
		
	def offset_elevations(self, table, offset, axis='Y'):
		
		for series in self.tab_file_data[table].keys():
			for idx, val in enumerate(self.tab_file_data[table][series][axis]):
				self.tab_file_data[table][series][axis][idx] = val + offset
				
	def get_first_series(self, table):
		series = list(self.tab_file_data[table].keys())[0]
		
		return series
		
if __name__ == "__main__":
	tab = TabFile(r"T:\114-Projects\OP211898 Maersk Developer Karoon\1.0 Weak Point Tether – 7-PRA-2-SPS\Manual Das File Change\-14pc Offset Survial 1-Yr Curr\analysis.tab")
	
	tab.export_to_excel(None)